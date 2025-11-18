"""
Comprehensive KB article download, clean, chunk, and ingest pipeline
Processes Excel file with KB links → downloads files → cleans data → creates atomic chunks
"""

import json
import os
import sys
import requests
from pathlib import Path
from datetime import datetime
import hashlib
import re
from typing import List, Dict

# Install required packages
required_packages = ['openpyxl', 'python-docx', 'PyPDF2', 'requests']
for package in required_packages:
    try:
        __import__(package.replace('-', '_'))
    except ImportError:
        print(f"Installing {package}...")
        os.system(f"{sys.executable} -m pip install {package} -q")

import openpyxl
import requests
from pathlib import Path

class KBProcessor:
    """Download and process KB articles"""
    
    def __init__(self, proj_root):
        self.proj_root = Path(proj_root)
        self.data_dir = self.proj_root / "data" / "kb_downloads"
        self.data_dir.mkdir(parents=True, exist_ok=True)
        self.excel_path = self.proj_root / "data" / "SOP_update_sheet.xlsx"
        self.links_json = self.proj_root / "data" / "kb_article_links.json"
        self.downloaded_dir = self.data_dir / "downloaded"
        self.downloaded_dir.mkdir(exist_ok=True)
        self.cleaned_dir = self.data_dir / "cleaned"
        self.cleaned_dir.mkdir(exist_ok=True)
    
    def parse_excel(self) -> List[Dict]:
        """Extract links from Excel file"""
        
        print(f"\n{'='*80}")
        print(f"[1/4] PARSING EXCEL FILE")
        print(f"{'='*80}\n")
        
        if not self.excel_path.exists():
            print(f"❌ Excel file not found: {self.excel_path}")
            print(f"\n📥 Please download from SharePoint and save to:")
            print(f"   {self.excel_path}")
            print(f"\nThen run this script again.\n")
            return []
        
        print(f"📖 Parsing: {self.excel_path.name}")
        
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            
            links = []
            print(f"📝 Scanning {ws.max_row} rows...\n")
            
            for row_idx, row in enumerate(ws.iter_rows(min_col=1, max_col=15, values_only=False), 1):
                for cell_idx, cell in enumerate(row, 1):
                    if cell.value:
                        value = str(cell.value).strip()
                        
                        # Match URLs
                        if value.startswith(('http://', 'https://', 'www.')):
                            title = f"KB_{len(links)+1:03d}"
                            
                            # Try to get title from nearby cells
                            for offset in range(-3, 4):
                                try:
                                    adj = ws.cell(row_idx, cell_idx + offset)
                                    if adj.value and not str(adj.value).startswith('http'):
                                        candidate = str(adj.value)[:80].strip()
                                        if candidate and len(candidate) > 3:
                                            title = candidate
                                            break
                                except:
                                    pass
                            
                            links.append({
                                "title": title,
                                "url": value,
                                "source": "sharepoint_excel",
                                "row": row_idx
                            })
                            
                            if len(links) % 10 == 0:
                                print(f"   Found {len(links)} links...", end='\r')
            
            print(f"\n✅ Extracted {len(links)} KB article links\n")
            
            # Save links
            with open(self.links_json, 'w', encoding='utf-8') as f:
                json.dump(links, f, indent=2, ensure_ascii=False)
            
            print(f"💾 Saved to: {self.links_json}\n")
            return links
        
        except Exception as e:
            print(f"❌ Error parsing Excel: {e}\n")
            return []
    
    def download_articles(self, links: List[Dict], max_articles: int = None) -> Dict:
        """Batch download KB articles"""
        
        print(f"{'='*80}")
        print(f"[2/4] DOWNLOADING KB ARTICLES")
        print(f"{'='*80}\n")
        
        if max_articles:
            links = links[:max_articles]
        
        results = {
            "total": len(links),
            "downloaded": 0,
            "failed": 0,
            "files": []
        }
        
        print(f"📥 Downloading {len(links)} articles...\n")
        
        for idx, link in enumerate(links, 1):
            url = link['url']
            title = link['title']
            
            try:
                print(f"   [{idx:3d}/{len(links)}] {title[:50]:<50}", end='... ', flush=True)
                
                # Download with timeout
                response = requests.get(url, timeout=10, allow_redirects=True)
                response.raise_for_status()
                
                # Determine file extension
                content_type = response.headers.get('content-type', '')
                if 'pdf' in content_type:
                    ext = '.pdf'
                elif 'word' in content_type or 'document' in content_type:
                    ext = '.docx'
                elif 'sheet' in content_type or 'excel' in content_type:
                    ext = '.xlsx'
                elif 'html' in content_type:
                    ext = '.html'
                else:
                    # Try from URL
                    if url.endswith('.pdf'):
                        ext = '.pdf'
                    elif url.endswith('.docx'):
                        ext = '.docx'
                    else:
                        ext = '.txt'
                
                # Save file
                safe_title = "".join(c for c in title if c.isalnum() or c in ' -_').strip()[:100]
                filename = f"{idx:03d}_{safe_title}{ext}"
                filepath = self.downloaded_dir / filename
                
                with open(filepath, 'wb') as f:
                    f.write(response.content)
                
                size_kb = os.path.getsize(filepath) / 1024
                print(f"✅ ({size_kb:.1f}KB)")
                
                results['downloaded'] += 1
                results['files'].append({
                    "title": title,
                    "url": url,
                    "file": str(filepath.relative_to(self.proj_root)),
                    "size_kb": round(size_kb, 1),
                    "downloaded": datetime.now().isoformat()
                })
            
            except Exception as e:
                print(f"❌ Error: {str(e)[:40]}")
                results['failed'] += 1
        
        print(f"\n{'='*80}")
        print(f"📊 DOWNLOAD SUMMARY")
        print(f"{'='*80}")
        print(f"Total articles:    {results['total']}")
        print(f"Downloaded:        {results['downloaded']} ✅")
        print(f"Failed:            {results['failed']} ❌")
        print(f"Success rate:      {results['downloaded']/results['total']*100:.1f}%\n")
        
        # Save results
        with open(self.data_dir / "download_results.json", 'w') as f:
            json.dump(results, f, indent=2)
        
        return results
    
    def clean_text(self, text: str) -> str:
        """Clean and normalize text"""
        # Remove extra whitespace
        text = re.sub(r'\s+', ' ', text)
        # Remove special characters but keep some structure
        text = re.sub(r'[\x00-\x08\x0b-\x0c\x0e-\x1f]', '', text)
        return text.strip()
    
    def extract_text_from_file(self, filepath: Path) -> str:
        """Extract text based on file type"""
        
        try:
            if filepath.suffix.lower() == '.pdf':
                try:
                    from PyPDF2 import PdfReader
                    with open(filepath, 'rb') as f:
                        reader = PdfReader(f)
                        text = ''.join(page.extract_text() for page in reader.pages)
                    return text
                except:
                    return ""
            
            elif filepath.suffix.lower() == '.docx':
                try:
                    from docx import Document
                    doc = Document(filepath)
                    text = '\n'.join(para.text for para in doc.paragraphs)
                    return text
                except:
                    return ""
            
            elif filepath.suffix.lower() in ['.txt', '.md']:
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    return f.read()
            
            elif filepath.suffix.lower() == '.html':
                try:
                    from html.parser import HTMLParser
                    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                        return f.read()
                except:
                    return ""
            
            return ""
        
        except Exception as e:
            print(f"Error extracting from {filepath}: {e}")
            return ""
    
    def process_downloads(self) -> Dict:
        """Clean and normalize downloaded files"""
        
        print(f"\n{'='*80}")
        print(f"[3/4] CLEANING & PREPROCESSING")
        print(f"{'='*80}\n")
        
        processed = {
            "total_files": 0,
            "processed": 0,
            "failed": 0,
            "total_tokens": 0,
            "articles": []
        }
        
        files = sorted(self.downloaded_dir.glob("*"))
        print(f"📂 Processing {len(files)} downloaded files...\n")
        
        for idx, filepath in enumerate(files, 1):
            print(f"   [{idx:3d}/{len(files)}] {filepath.name[:50]:<50}", end='... ', flush=True)
            
            try:
                # Extract text
                text = self.extract_text_from_file(filepath)
                
                if not text or len(text.strip()) < 50:
                    print(f"⚠️  Empty")
                    processed['failed'] += 1
                    continue
                
                # Clean text
                cleaned = self.clean_text(text)
                
                # Estimate tokens
                tokens = len(cleaned) // 4
                
                # Save cleaned version
                clean_filename = filepath.stem + ".txt"
                clean_filepath = self.cleaned_dir / clean_filename
                
                with open(clean_filepath, 'w', encoding='utf-8') as f:
                    f.write(cleaned)
                
                print(f"✅ ({tokens} tokens)")
                
                processed['processed'] += 1
                processed['total_tokens'] += tokens
                processed['articles'].append({
                    "original": filepath.name,
                    "cleaned": clean_filename,
                    "tokens": tokens,
                    "text_length": len(cleaned)
                })
            
            except Exception as e:
                print(f"❌ {str(e)[:30]}")
                processed['failed'] += 1
            
            processed['total_files'] = len(files)
        
        print(f"\n{'='*80}")
        print(f"📊 CLEANING SUMMARY")
        print(f"{'='*80}")
        print(f"Total files:       {processed['total_files']}")
        print(f"Processed:         {processed['processed']} ✅")
        print(f"Failed:            {processed['failed']} ❌")
        print(f"Total tokens:      {processed['total_tokens']:,}\n")
        
        # Save results
        with open(self.data_dir / "cleaning_results.json", 'w') as f:
            json.dump(processed, f, indent=2)
        
        return processed

def main():
    """Main pipeline"""
    
    proj_root = Path(__file__).parent.parent
    processor = KBProcessor(proj_root)
    
    # Step 1: Parse Excel
    links = processor.parse_excel()
    
    if not links:
        print("⚠️  No links found. Please ensure Excel file is in place and has article links.")
        return
    
    # Step 2: Download (limit to 10 for now to test)
    print("\n⏸️  Starting batch download (limiting to 10 articles for initial test)...")
    print("   (Remove max_articles parameter to download all)\n")
    
    results = processor.download_articles(links, max_articles=10)
    
    if results['downloaded'] == 0:
        print("⚠️  No articles downloaded. Check URLs and network connection.")
        return
    
    # Step 3: Clean downloaded files
    processed = processor.process_downloads()
    
    if processed['processed'] > 0:
        print("\n✅ PIPELINE COMPLETE!")
        print(f"\nNext steps:")
        print(f"1. Review cleaned files: {processor.cleaned_dir}")
        print(f"2. Run chunking: python scripts/chunk_kb_articles.py")
        print(f"3. Ingest into Chroma: python scripts/ingest_kb_chunks.py\n")

if __name__ == "__main__":
    main()
