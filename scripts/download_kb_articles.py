"""
Download Excel file from SharePoint and extract KB/SOP article links
"""

import os
import sys
import json
import requests
from pathlib import Path
from io import BytesIO

# Try to import openpyxl for Excel parsing
try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

def download_excel_from_sharepoint(url, output_path):
    """Download Excel file from SharePoint"""
    
    print(f"\n[1/3] Downloading Excel file from SharePoint...")
    print(f"      URL: {url}")
    
    try:
        # SharePoint direct download link
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        
        # Save file
        with open(output_path, 'wb') as f:
            f.write(response.content)
        
        file_size = os.path.getsize(output_path)
        print(f"      ✓ Downloaded: {file_size:,} bytes")
        print(f"      ✓ Saved to: {output_path}\n")
        
        return True
    except Exception as e:
        print(f"      ✗ Error: {e}")
        print(f"      Note: If this is a private SharePoint link, you may need to:")
        print(f"        1. Download manually & save to: {output_path}")
        print(f"        2. Then run step 2\n")
        return False

def extract_links_from_excel(excel_path, output_json):
    """Extract KB article links from Excel file"""
    
    print(f"[2/3] Parsing Excel file to extract KB/SOP links...")
    print(f"      File: {excel_path}")
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        # Extract all non-empty cells that look like URLs
        links = []
        titles = []
        
        print(f"      Scanning {ws.max_row} rows...\n")
        
        for row_idx, row in enumerate(ws.iter_rows(min_col=1, max_col=10, values_only=False), 1):
            for cell_idx, cell in enumerate(row, 1):
                if cell.value:
                    value = str(cell.value).strip()
                    
                    # Check if it's a URL
                    if value.startswith('http'):
                        # Extract title from adjacent cells if available
                        title = f"Article_{len(links)+1}"
                        if row_idx > 0:
                            for offset in range(-2, 3):
                                try:
                                    adj_cell = ws.cell(row_idx, cell_idx + offset)
                                    if adj_cell.value and not str(adj_cell.value).startswith('http'):
                                        title = str(adj_cell.value)[:100]
                                        break
                                except:
                                    pass
                        
                        links.append({
                            "title": title,
                            "url": value,
                            "row": row_idx,
                            "col": cell_idx
                        })
                        
                        print(f"      [{len(links):3d}] {title[:50]:<50} | {value[:60]}")
        
        print(f"\n      ✓ Found {len(links)} links\n")
        
        # Save to JSON
        with open(output_json, 'w', encoding='utf-8') as f:
            json.dump(links, f, indent=2, ensure_ascii=False)
        
        print(f"      ✓ Saved to: {output_json}\n")
        
        return links
    
    except Exception as e:
        print(f"      ✗ Error parsing Excel: {e}")
        print(f"      Make sure the file is valid and try again\n")
        return []

def main():
    """Download and parse KB links"""
    
    proj_root = Path(__file__).parent
    excel_url = "https://cloudspacetechnologies-my.sharepoint.com/:x:/r/personal/sumant_choudhary_myrealdata_in/Documents/SOP%20update%20sheet.xlsx?d=wd193acb3653747019df91d686db6fa01&csf=1&web=1&e=44U5Ec"
    
    excel_path = proj_root / "data" / "SOP_update_sheet.xlsx"
    links_json = proj_root / "data" / "kb_article_links.json"
    
    os.makedirs(proj_root / "data", exist_ok=True)
    
    print("\n" + "="*80)
    print("KB ARTICLE DOWNLOADER - STEP 1: Extract Links from Excel")
    print("="*80)
    
    # Step 1: Download Excel
    if not excel_path.exists():
        success = download_excel_from_sharepoint(excel_url, excel_path)
        if not success:
            print("⚠ Could not auto-download. Please download manually and place at:")
            print(f"  {excel_path}")
            print("\nThen run: python scripts/download_kb_articles.py --step2")
            return
    else:
        print(f"\n✓ Excel file already exists: {excel_path}\n")
    
    # Step 2: Extract links
    links = extract_links_from_excel(excel_path, links_json)
    
    if links:
        print("="*80)
        print(f"SUCCESS! Found {len(links)} KB article links")
        print("="*80)
        print(f"\nNext: Run batch downloader to fetch all articles")
        print(f"  python scripts/download_kb_articles.py --step2\n")
    else:
        print("⚠ No links found. Check Excel file format and try again.\n")

if __name__ == "__main__":
    main()
